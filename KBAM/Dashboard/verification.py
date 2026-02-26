import pandas as pd
from call import get_strategy_portfolio_weight_comparison
from typing import Optional
import sys


def render_verification(index_name: str, base_date: str, end_date: str):
    """
    전략 포트폴리오 비중 검증 섹션 렌더링 (Streamlit용)
    절대 금액(NAV) 기준으로 먼저 표시하고, 비중은 보조 정보로 제공
    
    Args:
        index_name: 지수명 (BM)
        base_date: 기준일자 (YYYY-MM-DD 형식)
        end_date: 종료일자 (YYYY-MM-DD 형식)
    """
    import streamlit as st
    from io import BytesIO
    
    st.markdown("---")
    st.subheader("📊 전략 포트폴리오 비중 검증")
    
    with st.spinner("전략 포트폴리오 비중 비교 데이터를 생성하는 중..."):
        weight_comparison_data = get_strategy_portfolio_weight_comparison(
            index_name=index_name,
            base_date=base_date,
            end_date=end_date
        )
        
        if not weight_comparison_data.empty:
            # 날짜별 요약 정보 가져오기
            daily_weight_summary = None
            if hasattr(weight_comparison_data, 'attrs') and 'daily_weight_summary' in weight_comparison_data.attrs:
                daily_weight_summary = weight_comparison_data.attrs['daily_weight_summary']
            
            # ============================================
            # ① 포트폴리오 전체 요약 (맨 위)
            # ============================================
            st.markdown("### ① 포트폴리오 전체 요약")
            
            if daily_weight_summary is not None and not daily_weight_summary.empty:
                # NAV 컬럼이 있는지 확인
                if all(col in daily_weight_summary.columns for col in ['BM_NAV', 'MP_NAV', 'NAV_차이']):
                    # 가장 최근 날짜의 NAV 정보 표시
                    latest_summary = daily_weight_summary.iloc[-1]
                    
                    nav_summary_data = {
                        '항목': ['BM NAV', 'MP NAV', 'NAV 차이'],
                        '의미': [
                            '기준 포트 총 자산',
                            '실제 운용 포트 총 자산 (추가투입 포함)',
                            'MP − BM (추가투입 금액)'
                        ],
                        '값': [
                            latest_summary.get('BM_NAV', 0),
                            latest_summary.get('MP_NAV', 0),
                            latest_summary.get('NAV_차이', 0)
                        ]
                    }
                    
                    nav_summary_df = pd.DataFrame(nav_summary_data)
                    st.dataframe(
                        nav_summary_df.style.format({
                            '값': '{:,.4f}'
                        }),
                        use_container_width=True,
                        hide_index=True
                    )
                    
                    st.caption("📌 **1% 추가투입 = NAV가 커진다**는 게 바로 보임")
                else:
                    st.info("NAV 정보를 사용할 수 없습니다. 데이터를 다시 생성해주세요.")
                st.markdown("---")
            
            # ============================================
            # ② 종목별 상세 (핵심 테이블) - 절대 금액 기준
            # ============================================
            st.markdown("### ② 종목별 상세 (절대 금액 기준)")
            st.caption("👉 **비중이 아니라 실제 돈 기준으로 먼저 보여라**")
            
            # 가장 최근 날짜의 데이터만 필터링
            if '날짜' in weight_comparison_data.columns:
                latest_date = weight_comparison_data['날짜'].max()
                latest_data = weight_comparison_data[weight_comparison_data['날짜'] == latest_date].copy()
            else:
                latest_data = weight_comparison_data.copy()
            
            # 절대 금액 기준 테이블 생성
            if 'BM_금액' in latest_data.columns and 'MP_금액' in latest_data.columns:
                absolute_table = latest_data[[
                    '종목명', 'BM_금액', 'MP_금액', '절대_Active_금액', '절대_Active_비율'
                ]].copy()
                
                # 컬럼명 변경
                absolute_table.columns = ['종목', 'BM 금액', 'MP 금액', '절대 Active (₩)', '절대 Active (%)']
                
                # 정렬: 절대 Active 금액이 큰 순서대로
                absolute_table = absolute_table.sort_values('절대 Active (₩)', ascending=False)
                
                # 표시용 포맷팅
                absolute_table_display = absolute_table.copy()
                absolute_table_display['절대 Active (%)'] = absolute_table_display['절대 Active (%)'] * 100
                
                st.dataframe(
                    absolute_table_display.style.format({
                        'BM 금액': '{:,.4f}',
                        'MP 금액': '{:,.4f}',
                        '절대 Active (₩)': '{:,.4f}',
                        '절대 Active (%)': '{:.2f}%'
                    }),
                    use_container_width=True,
                    hide_index=True
                )
                
                st.caption("📌 **이 표만 봐도 \"A는 진짜로 1% 더 샀다\"가 명확**")
                st.caption("📌 **이 단계에서는 정규화 금지**")
                st.markdown("---")
            
            # ============================================
            # ③ 참고용: 정규화된 비중 (보조 정보)
            # ============================================
            st.markdown("### ③ 참고용: 정규화된 비중")
            st.caption("👉 **비중 관점에서 보면 이렇게 보인다**를 추가로 보여줌")
            
            if 'BM_비중' in latest_data.columns and '전략_비중' in latest_data.columns:
                weight_table = latest_data[[
                    '종목명', 'BM_비중', '전략_비중', '비중_차이'
                ]].copy()
                
                # 정규화된 MP Weight 계산 (MP NAV 기준으로 정규화)
                if daily_weight_summary is not None and not daily_weight_summary.empty and 'MP_NAV' in daily_weight_summary.columns:
                    latest_summary = daily_weight_summary.iloc[-1]
                    mp_nav = latest_summary.get('MP_NAV', 1.0)
                    if mp_nav > 0:
                        # MP 금액을 MP NAV로 나누어 정규화된 비중 계산
                        if 'MP_금액' in latest_data.columns:
                            weight_table['MP_Weight_정규화'] = latest_data['MP_금액'] / mp_nav
                        else:
                            weight_table['MP_Weight_정규화'] = weight_table['전략_비중']
                    else:
                        weight_table['MP_Weight_정규화'] = weight_table['전략_비중']
                else:
                    weight_table['MP_Weight_정규화'] = weight_table['전략_비중']
                
                # Weight 차이 (정규화된 기준)
                weight_table['Weight_차이_정규화'] = weight_table['MP_Weight_정규화'] - weight_table['BM_비중']
                
                # 컬럼명 변경
                weight_table = weight_table[[
                    '종목명', 'BM_비중', 'MP_Weight_정규화', 'Weight_차이_정규화'
                ]]
                weight_table.columns = ['종목', 'BM Weight', 'MP Weight (정규화)', 'Weight 차이']
                
                # 정렬: Weight 차이가 큰 순서대로
                weight_table = weight_table.sort_values('Weight 차이', ascending=False)
                
                # 표시용 포맷팅
                weight_table_display = weight_table.copy()
                weight_table_display['BM Weight'] = weight_table_display['BM Weight'] * 100
                weight_table_display['MP Weight (정규화)'] = weight_table_display['MP Weight (정규화)'] * 100
                weight_table_display['Weight 차이'] = weight_table_display['Weight 차이'] * 100
                
                st.dataframe(
                    weight_table_display.style.format({
                        'BM Weight': '{:.2f}%',
                        'MP Weight (정규화)': '{:.2f}%',
                        'Weight 차이': '{:.2f}%'
                    }),
                    use_container_width=True,
                    hide_index=True
                )
                
                st.caption("📌 **여기서 1%가 0.4%처럼 보이는 현상 발생**")
                st.caption("📌 **하지만 위 ②번 표가 있어서 왜곡이 아님**")
                st.markdown("---")
            
            # ============================================
            # ④ 일별 변화 (차트 or 보조 테이블)
            # ============================================
            st.markdown("### ④ 일별 변화")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("#### (a) 절대 기준")
                st.caption("종목별 MP 금액 변화, Active 금액 변화")
                
                if daily_weight_summary is not None and not daily_weight_summary.empty:
                    # NAV 컬럼이 있는지 확인
                    if all(col in daily_weight_summary.columns for col in ['BM_NAV', 'MP_NAV', 'NAV_차이']):
                        # NAV 변화 테이블
                        nav_change_table = daily_weight_summary[[
                            '날짜', 'BM_NAV', 'MP_NAV', 'NAV_차이'
                        ]].copy()
                        nav_change_table.columns = ['날짜', 'BM NAV', 'MP NAV', 'NAV 차이']
                        
                        st.dataframe(
                            nav_change_table.style.format({
                                'BM NAV': '{:,.4f}',
                                'MP NAV': '{:,.4f}',
                                'NAV 차이': '{:,.4f}'
                            }),
                            use_container_width=True,
                            hide_index=True
                        )
                        
                        # 차트로도 표시
                        if len(nav_change_table) > 1:
                            chart_data = nav_change_table.set_index('날짜')[['BM NAV', 'MP NAV']]
                            st.line_chart(chart_data)
                    else:
                        st.info("NAV 정보를 사용할 수 없습니다.")
            
            with col2:
                st.markdown("#### (b) 비중 기준")
                st.caption("정규화된 MP weight 변화")
                
                if daily_weight_summary is not None and not daily_weight_summary.empty:
                    # 비중 변화 테이블
                    weight_change_table = daily_weight_summary[[
                        '날짜', 'BM_비중_합계', '전략_비중_합계', '비중_합계_차이'
                    ]].copy()
                    weight_change_table.columns = ['날짜', 'BM Weight', 'MP Weight', 'Weight 차이']
                    
                    # 표시용 포맷팅
                    weight_change_table_display = weight_change_table.copy()
                    weight_change_table_display['BM Weight'] = weight_change_table_display['BM Weight'] * 100
                    weight_change_table_display['MP Weight'] = weight_change_table_display['MP Weight'] * 100
                    weight_change_table_display['Weight 차이'] = weight_change_table_display['Weight 차이'] * 100
                    
                    st.dataframe(
                        weight_change_table_display.style.format({
                            'BM Weight': '{:.2f}%',
                            'MP Weight': '{:.2f}%',
                            'Weight 차이': '{:.2f}%'
                        }),
                        use_container_width=True,
                        hide_index=True
                    )
                    
                    # 차트로도 표시
                    if len(weight_change_table_display) > 1:
                        chart_data = weight_change_table_display.set_index('날짜')[['BM Weight', 'MP Weight']]
                        st.line_chart(chart_data)
            
            st.markdown("---")
            
            # ============================================
            # ⑤ 성과 요약 (맨 아래)
            # ============================================
            st.markdown("### ⑤ 성과 요약")
            st.caption("👉 **결과는 이렇게 단순하게**")
            
            # BM return과 MP return 계산
            # 기준일자 대비 수익률을 종목별로 계산하고, 비중 가중 평균
            if '기준일자_대비_수익률' in weight_comparison_data.columns:
                # 기준일자 데이터
                base_date_data = weight_comparison_data[weight_comparison_data['날짜'] == base_date].copy()
                if base_date_data.empty:
                    # 기준일자가 없으면 첫 번째 날짜 사용
                    first_date = weight_comparison_data['날짜'].min()
                    base_date_data = weight_comparison_data[weight_comparison_data['날짜'] == first_date].copy()
                
                # 가장 최근 날짜 데이터
                latest_date = weight_comparison_data['날짜'].max()
                latest_perf_data = weight_comparison_data[weight_comparison_data['날짜'] == latest_date].copy()
                
                if not base_date_data.empty and not latest_perf_data.empty:
                    # BM return = BM 비중 * 수익률의 합
                    bm_return = (latest_perf_data['BM_비중'] * latest_perf_data['기준일자_대비_수익률']).sum() * 100
                    
                    # MP return = 전략 비중 * 수익률의 합
                    mp_return = (latest_perf_data['전략_비중'] * latest_perf_data['기준일자_대비_수익률']).sum() * 100
                    
                    # Absolute Alpha = MP return - BM return (금액 기준)
                    # NAV 기준으로 계산
                    if daily_weight_summary is not None and not daily_weight_summary.empty and 'BM_NAV' in daily_weight_summary.columns:
                        latest_summary = daily_weight_summary.iloc[-1]
                        bm_nav = latest_summary.get('BM_NAV', 1.0)
                        absolute_alpha = (mp_return - bm_return) / 100 * bm_nav
                    else:
                        absolute_alpha = (mp_return - bm_return) / 100
                    
                    # Relative Alpha = MP return - BM return (%)
                    relative_alpha = mp_return - bm_return
                    
                    performance_summary = {
                        '항목': ['BM return', 'MP return', 'Absolute Alpha', 'Relative Alpha'],
                        '의미': [
                            '기준 성과',
                            '실제 성과',
                            'MP − BM (금액)',
                            'MP − BM (%)'
                        ],
                        '값': [bm_return, mp_return, absolute_alpha, relative_alpha]
                    }
                    
                    perf_summary_df = pd.DataFrame(performance_summary)
                    
                    # 표시용 데이터프레임 생성
                    perf_summary_display = perf_summary_df.copy()
                    perf_summary_display['값_표시'] = perf_summary_display.apply(
                        lambda row: f'{row["값"]:,.4f}' if row['항목'] == 'Absolute Alpha' else f'{row["값"]:.2f}%',
                        axis=1
                    )
                    perf_summary_display = perf_summary_display[['항목', '의미', '값_표시']]
                    perf_summary_display.columns = ['항목', '의미', '값']
                    
                    st.dataframe(
                        perf_summary_display,
                        use_container_width=True,
                        hide_index=True
                    )
            
            st.markdown("---")
            
            # ============================================
            # 전체 데이터 다운로드 (기존 기능 유지)
            # ============================================
            st.markdown("#### 전체 데이터 다운로드")
            
            # 엑셀 다운로드 버튼
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # ============================================
                # ① 일별 포트 수익률 (핵심 KPI)
                # ============================================
                if daily_weight_summary is not None and not daily_weight_summary.empty:
                    if all(col in daily_weight_summary.columns for col in ['BM_NAV', 'MP_NAV']):
                        # 날짜별로 정렬
                        daily_weight_sorted = daily_weight_summary.sort_values('날짜').copy()
                        
                        daily_returns = []
                        prev_bm_nav = None
                        prev_mp_nav = None
                        
                        for _, row in daily_weight_sorted.iterrows():
                            date = row['날짜']
                            bm_nav = row.get('BM_NAV', 0)
                            mp_nav = row.get('MP_NAV', 0)
                            
                            # 일별 수익률 계산 (전일 대비)
                            if prev_bm_nav is not None and prev_bm_nav > 0:
                                bm_daily_return = ((bm_nav / prev_bm_nav) - 1) * 100
                            else:
                                bm_daily_return = 0.0
                            
                            if prev_mp_nav is not None and prev_mp_nav > 0:
                                mp_daily_return = ((mp_nav / prev_mp_nav) - 1) * 100
                            else:
                                mp_daily_return = 0.0
                            
                            daily_alpha = mp_daily_return - bm_daily_return
                            
                            daily_returns.append({
                                'Date': date,
                                'BM Return': f'{bm_daily_return:.2f}%',
                                'MP Return': f'{mp_daily_return:.2f}%',
                                'Daily Alpha': f'{daily_alpha:.2f}%'
                            })
                            
                            prev_bm_nav = bm_nav
                            prev_mp_nav = mp_nav
                        
                        if daily_returns:
                            daily_returns_df = pd.DataFrame(daily_returns)
                            daily_returns_df.to_excel(
                                writer,
                                sheet_name='①_일별_포트수익률',
                                index=False
                            )
                
                # ============================================
                # ② 누적 NAV 추이 (대시보드 메인 차트)
                # ============================================
                if daily_weight_summary is not None and not daily_weight_summary.empty:
                    if all(col in daily_weight_summary.columns for col in ['BM_NAV', 'MP_NAV']):
                        # Start 행 추가 (기준일자)
                        base_date_summary = daily_weight_summary[daily_weight_summary['날짜'] == base_date]
                        if base_date_summary.empty:
                            first_date = daily_weight_summary['날짜'].min()
                            base_date_summary = daily_weight_summary[daily_weight_summary['날짜'] == first_date]
                        
                        nav_trend = daily_weight_summary[['날짜', 'BM_NAV', 'MP_NAV']].copy()
                        nav_trend.columns = ['Date', 'BM NAV', 'MP NAV']
                        
                        # Start 행 추가
                        if not base_date_summary.empty:
                            start_row = base_date_summary.iloc[0]
                            start_df = pd.DataFrame({
                                'Date': ['Start'],
                                'BM NAV': [start_row.get('BM_NAV', 0)],
                                'MP NAV': [start_row.get('MP_NAV', 0)]
                            })
                            nav_trend = pd.concat([start_df, nav_trend], ignore_index=True)
                        
                        nav_trend.to_excel(
                            writer,
                            sheet_name='②_누적_NAV_추이',
                            index=False
                        )
                
                # ============================================
                # ③ Active 포지션 모니터링 (절대 기준)
                # ============================================
                if '날짜' in weight_comparison_data.columns and '절대_Active_금액' in weight_comparison_data.columns:
                    # Active 금액이 있는 종목만 필터링
                    active_stocks = weight_comparison_data[weight_comparison_data['절대_Active_금액'] != 0].copy()
                    
                    if not active_stocks.empty:
                        # 각 종목별로 별도 테이블 생성
                        for stock_name in active_stocks['종목명'].unique():
                            stock_data = active_stocks[active_stocks['종목명'] == stock_name].sort_values('날짜')
                            
                            active_monitoring = []
                            prev_active_amount = None
                            prev_price = None
                            total_pnl = 0.0
                            
                            # Start 행 추가 (기준일자)
                            base_date_data = stock_data[stock_data['날짜'] == base_date]
                            if base_date_data.empty:
                                first_date = stock_data['날짜'].min()
                                base_date_data = stock_data[stock_data['날짜'] == first_date]
                            
                            if not base_date_data.empty:
                                start_row = base_date_data.iloc[0]
                                start_active_amount = start_row.get('절대_Active_금액', 0)
                                start_price = start_row.get('PRICE', None)
                                
                                active_monitoring.append({
                                    'Date': 'Start',
                                    f'{stock_name} Active Amount': start_active_amount,
                                    f'{stock_name} Return': '–',
                                    f'{stock_name} Active P&L': 0.00
                                })
                                
                                prev_active_amount = start_active_amount
                                prev_price = start_price
                            
                            # 일별 데이터
                            for _, row in stock_data.iterrows():
                                date = row['날짜']
                                active_amount = row.get('절대_Active_금액', 0)
                                current_price = row.get('PRICE', None)
                                
                                # 일별 수익률 계산
                                if prev_price is not None and prev_price > 0 and current_price is not None:
                                    daily_return = ((current_price / prev_price) - 1) * 100
                                    # Active P&L = 전일 Active Amount × 일별 수익률
                                    active_pnl = prev_active_amount * (daily_return / 100)
                                    total_pnl += active_pnl
                                else:
                                    daily_return = 0.0
                                    active_pnl = 0.0
                                
                                active_monitoring.append({
                                    'Date': date,
                                    f'{stock_name} Active Amount': active_amount,
                                    f'{stock_name} Return': f'{daily_return:.1f}%' if daily_return != 0 else '–',
                                    f'{stock_name} Active P&L': f'{active_pnl:.3f}'
                                })
                                
                                prev_active_amount = active_amount
                                prev_price = current_price
                            
                            # 합계 행 추가
                            active_monitoring.append({
                                'Date': '합계',
                                f'{stock_name} Active Amount': '',
                                f'{stock_name} Return': '',
                                f'{stock_name} Active P&L': f'{total_pnl:.3f}'
                            })
                            
                            if active_monitoring:
                                active_df = pd.DataFrame(active_monitoring)
                                sheet_name = f'③_Active_{stock_name}' if len(active_stocks['종목명'].unique()) > 1 else '③_Active_포지션_모니터링'
                                active_df.to_excel(
                                    writer,
                                    sheet_name=sheet_name,
                                    index=False
                                )
                
                # ============================================
                # ④ 참고용: 정규화된 비중 (보조 차트)
                # ============================================
                if '날짜' in weight_comparison_data.columns:
                    # Active 금액이 있는 종목만 선택
                    active_stocks = weight_comparison_data[weight_comparison_data['절대_Active_금액'] != 0].copy()
                    
                    if not active_stocks.empty and daily_weight_summary is not None and not daily_weight_summary.empty:
                        # 각 종목별로 별도 테이블 생성
                        for stock_name in active_stocks['종목명'].unique():
                            stock_data = active_stocks[active_stocks['종목명'] == stock_name].sort_values('날짜')
                            
                            normalized_weights = []
                            
                            # Start 행 추가
                            base_date_data = stock_data[stock_data['날짜'] == base_date]
                            if base_date_data.empty:
                                first_date = stock_data['날짜'].min()
                                base_date_data = stock_data[stock_data['날짜'] == first_date]
                            
                            if not base_date_data.empty:
                                start_row = base_date_data.iloc[0]
                                bm_weight = start_row.get('BM_비중', 0) * 100
                                
                                # Start일의 MP Weight 계산
                                start_date_summary = daily_weight_summary[daily_weight_summary['날짜'] == start_row['날짜']]
                                if not start_date_summary.empty and 'MP_NAV' in start_date_summary.columns:
                                    mp_nav = start_date_summary.iloc[0].get('MP_NAV', 1.0)
                                    mp_amount = start_row.get('MP_금액', 0)
                                    if mp_nav > 0:
                                        mp_weight_normalized = (mp_amount / mp_nav) * 100
                                    else:
                                        mp_weight_normalized = bm_weight
                                else:
                                    mp_weight_normalized = bm_weight
                                
                                weight_diff = mp_weight_normalized - bm_weight
                                
                                normalized_weights.append({
                                    'Date': 'Start',
                                    f'{stock_name} Weight (MP, %)': f'{mp_weight_normalized:.2f}%',
                                    f'{stock_name} BM 대비': f'{weight_diff:.2f}%'
                                })
                            
                            # 일별 데이터 (주요 날짜만 선택 - Start, 중간, 최종)
                            dates_sorted = sorted(stock_data['날짜'].unique())
                            # Start, 중간 1개, 최종만 선택
                            if len(dates_sorted) > 2:
                                selected_dates = [dates_sorted[0], dates_sorted[len(dates_sorted)//2], dates_sorted[-1]]
                            else:
                                selected_dates = dates_sorted
                            
                            for date in selected_dates:
                                if date == base_date or date == dates_sorted[0]:
                                    continue  # Start는 이미 추가됨
                                
                                row = stock_data[stock_data['날짜'] == date].iloc[0]
                                bm_weight = row.get('BM_비중', 0) * 100
                                
                                # MP Weight (정규화) 계산
                                date_summary = daily_weight_summary[daily_weight_summary['날짜'] == date]
                                if not date_summary.empty and 'MP_NAV' in date_summary.columns:
                                    mp_nav = date_summary.iloc[0].get('MP_NAV', 1.0)
                                    mp_amount = row.get('MP_금액', 0)
                                    if mp_nav > 0:
                                        mp_weight_normalized = (mp_amount / mp_nav) * 100
                                    else:
                                        mp_weight_normalized = bm_weight
                                else:
                                    mp_weight_normalized = bm_weight
                                
                                weight_diff = mp_weight_normalized - bm_weight
                                
                                normalized_weights.append({
                                    'Date': date,
                                    f'{stock_name} Weight (MP, %)': f'{mp_weight_normalized:.2f}%',
                                    f'{stock_name} BM 대비': f'{weight_diff:.2f}%'
                                })
                            
                            if normalized_weights:
                                normalized_df = pd.DataFrame(normalized_weights)
                                sheet_name = f'④_정규화비중_{stock_name}' if len(active_stocks['종목명'].unique()) > 1 else '④_참고용_정규화된_비중'
                                normalized_df.to_excel(
                                    writer,
                                    sheet_name=sheet_name,
                                    index=False
                                )
                                
                                # 주석 추가
                                from openpyxl.comments import Comment
                                ws = writer.sheets[sheet_name]
                                # 첫 번째 데이터 행에 주석 추가
                                ws.cell(row=2, column=1).comment = Comment(
                                    "MP는 101% 포트이며, 본 비중은 정규화된 참고값",
                                    "시스템"
                                )
                            from openpyxl.comments import Comment
                            ws = writer.sheets['④_참고용_정규화된_비중']
                            # 첫 번째 데이터 행에 주석 추가
                            note_cell = ws.cell(row=2, column=1)
                            note_cell.comment = Comment(
                                "MP는 101% 포트이며, 본 비중은 정규화된 참고값",
                                "시스템"
                            )
                
                
                # ============================================
                # ⑤ 성과 요약 (임원/고객용)
                # ============================================
                if '기준일자_대비_수익률' in weight_comparison_data.columns:
                    base_date_data = weight_comparison_data[weight_comparison_data['날짜'] == base_date].copy()
                    if base_date_data.empty:
                        first_date = weight_comparison_data['날짜'].min()
                        base_date_data = weight_comparison_data[weight_comparison_data['날짜'] == first_date].copy()
                    
                    latest_date = weight_comparison_data['날짜'].max()
                    latest_perf_data = weight_comparison_data[weight_comparison_data['날짜'] == latest_date].copy()
                    
                    if not base_date_data.empty and not latest_perf_data.empty:
                        # BM 누적 수익률 = 기준일자 대비 수익률
                        bm_return = (latest_perf_data['BM_비중'] * latest_perf_data['기준일자_대비_수익률']).sum() * 100
                        
                        # MP 누적 수익률 = Σ (MP_amount × 종목수익률) / MP_NAV
                        if daily_weight_summary is not None and not daily_weight_summary.empty and 'MP_NAV' in daily_weight_summary.columns:
                            latest_summary = daily_weight_summary.iloc[-1]
                            mp_nav = latest_summary.get('MP_NAV', 1.0)
                            if mp_nav > 0 and 'MP_금액' in latest_perf_data.columns:
                                # MP_amount × 종목수익률의 합
                                mp_total_return = (latest_perf_data['MP_금액'] * latest_perf_data['기준일자_대비_수익률']).sum() * 100
                                mp_return = mp_total_return / mp_nav
                            else:
                                mp_return = (latest_perf_data['전략_비중'] * latest_perf_data['기준일자_대비_수익률']).sum() * 100
                        else:
                            mp_return = (latest_perf_data['전략_비중'] * latest_perf_data['기준일자_대비_수익률']).sum() * 100
                        
                        # Relative Alpha (%) = MP_return - BM_return
                        relative_alpha = mp_return - bm_return
                        
                        # Absolute Alpha (₩) = MP_NAV × (MP_return - BM_return) / 100
                        if daily_weight_summary is not None and not daily_weight_summary.empty and 'MP_NAV' in daily_weight_summary.columns:
                            latest_summary = daily_weight_summary.iloc[-1]
                            mp_nav = latest_summary.get('MP_NAV', 1.0)
                            absolute_alpha = mp_nav * (mp_return - bm_return) / 100
                        else:
                            absolute_alpha = (mp_return - bm_return) / 100
                        
                        # Alpha Source: Active 금액이 있는 종목들
                        active_stocks_list = []
                        if '절대_Active_금액' in latest_perf_data.columns:
                            active_stocks_data = latest_perf_data[latest_perf_data['절대_Active_금액'] != 0]
                            if not active_stocks_data.empty:
                                for _, row in active_stocks_data.iterrows():
                                    stock_name = row['종목명']
                                    active_pct = row.get('절대_Active_비율', 0) * 100
                                    if active_pct > 0:
                                        active_stocks_list.append(f"{stock_name} {active_pct:.1f}% OW")
                                    elif active_pct < 0:
                                        active_stocks_list.append(f"{stock_name} {abs(active_pct):.1f}% UW")
                        
                        alpha_source = ", ".join(active_stocks_list) if active_stocks_list else "없음"
                        
                        performance_summary = pd.DataFrame({
                            '항목': ['BM 누적 수익률', 'MP 누적 수익률', 'Relative Alpha', 'Absolute Alpha', 'Alpha Source'],
                            '값': [
                                f'{bm_return:.2f}%',
                                f'{mp_return:.2f}%',
                                f'{relative_alpha:.2f}%',
                                f'{absolute_alpha:.4f}',
                                alpha_source
                            ]
                        })
                        performance_summary.to_excel(
                            writer,
                            sheet_name='⑤_성과_요약',
                            index=False
                        )
                
                # 전체 데이터는 제거 (핵심 정보만 제공)
            
            output.seek(0)
            
            st.download_button(
                label="📥 전략 포트폴리오 비중 비교 데이터 다운로드 (Excel)",
                data=output,
                file_name=f"전략포트폴리오_비중비교_{index_name}_{base_date.replace('-', '')}_{end_date.replace('-', '')}.xlsx",
                mime="application/vnd.openpyxl-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("전략 포트폴리오 비중 비교 데이터를 생성할 수 없습니다.")


def save_verification_excel(index_name: str, base_date: str, end_date: str, output_path: Optional[str] = None):
    """
    전략 포트폴리오 비중 비교 데이터를 엑셀 파일로 저장하는 함수
    (Streamlit 없이 독립 실행 가능)
    
    Args:
        index_name: 지수명 (BM)
        base_date: 기준일자 (YYYY-MM-DD 형식)
        end_date: 종료일자 (YYYY-MM-DD 형식)
        output_path: 저장할 파일 경로 (None이면 자동 생성)
    
    Returns:
        str: 저장된 파일 경로
    """
    weight_comparison_data = get_strategy_portfolio_weight_comparison(
        index_name=index_name,
        base_date=base_date,
        end_date=end_date
    )
    
    if weight_comparison_data.empty:
        print("전략 포트폴리오 비중 비교 데이터를 생성할 수 없습니다.")
        return None
    
    print(f"데이터 로드 완료: {len(weight_comparison_data)}건")
    
    # 출력 파일 경로 설정
    import os
    # 스크립트 파일이 있는 디렉토리의 output 폴더
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        print(f"스크립트 디렉토리: {script_dir}")
    except NameError:
        # __file__이 없는 경우 (예: 인터랙티브 모드) 현재 작업 디렉토리 사용
        script_dir = os.getcwd()
        print(f"현재 작업 디렉토리 사용: {script_dir}")
    
    output_dir = os.path.join(script_dir, 'output')
    output_dir = os.path.abspath(output_dir)  # 절대 경로로 변환
    
    # output 폴더가 없으면 생성
    try:
        os.makedirs(output_dir, exist_ok=True)
        if not os.path.exists(output_dir):
            raise Exception(f"폴더 생성 실패: {output_dir}")
        print(f"저장 디렉토리: {output_dir}")
    except Exception as e:
        print(f"오류: output 폴더 생성 실패 - {e}")
        raise
    
    if output_path is None:
        filename = f"전략포트폴리오_비중비교_{index_name}_{base_date.replace('-', '')}_{end_date.replace('-', '')}.xlsx"
        output_path = os.path.join(output_dir, filename)
    else:
        # 절대 경로가 아니면 output 디렉토리 기준으로 변환
        if not os.path.isabs(output_path):
            output_path = os.path.join(output_dir, output_path)
    
    output_path = os.path.abspath(output_path)  # 절대 경로로 변환
    print(f"파일 저장 경로: {output_path}")
    
    # 날짜별 비중 합계 비교
    daily_weight_summary = None
    if hasattr(weight_comparison_data, 'attrs') and 'daily_weight_summary' in weight_comparison_data.attrs:
        daily_weight_summary = weight_comparison_data.attrs['daily_weight_summary']
    
    # 엑셀 파일 저장
    print(f"엑셀 파일 저장 시작...")
    try:
        # 기존 파일이 있으면 삭제
        if os.path.exists(output_path):
            os.remove(output_path)
            print(f"기존 파일 삭제: {output_path}")
        
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        
        # ============================================
        # ① 일별 포트 수익률 (핵심 KPI)
        # ============================================
        if daily_weight_summary is not None and not daily_weight_summary.empty and '날짜' in weight_comparison_data.columns:
            print(f"①_일별_포트수익률 시트 작성 중...")
            # 날짜별로 그룹화하여 일별 수익률 계산
            daily_returns = []
            dates_sorted = sorted(daily_weight_summary['날짜'].unique())
            
            prev_bm_nav = None
            prev_mp_nav = None
            
            for date in dates_sorted:
                date_summary = daily_weight_summary[daily_weight_summary['날짜'] == date]
                
                if not date_summary.empty:
                    date_summary_row = date_summary.iloc[0]
                    bm_nav = date_summary_row.get('BM_NAV', 0)
                    mp_nav = date_summary_row.get('MP_NAV', 0)
                    
                    # 일별 수익률 계산
                    if prev_bm_nav is not None and prev_bm_nav > 0:
                        bm_daily_return = ((bm_nav / prev_bm_nav) - 1) * 100
                    else:
                        bm_daily_return = 0.0
                    
                    if prev_mp_nav is not None and prev_mp_nav > 0:
                        mp_daily_return = ((mp_nav / prev_mp_nav) - 1) * 100
                    else:
                        mp_daily_return = 0.0
                    
                    daily_alpha = mp_daily_return - bm_daily_return
                    
                    daily_returns.append({
                        'Date': date,
                        'BM Return': f'{bm_daily_return:.2f}%',
                        'MP Return': f'{mp_daily_return:.2f}%',
                        'Daily Alpha': f'{daily_alpha:.2f}%'
                    })
                    
                    prev_bm_nav = bm_nav
                    prev_mp_nav = mp_nav
            
            if daily_returns:
                daily_returns_df = pd.DataFrame(daily_returns)
                daily_returns_df.to_excel(
                    writer,
                    sheet_name='①_일별_포트수익률',
                    index=False
                )
        
        # ============================================
        # ② 누적 NAV 추이 (대시보드 메인 차트)
        # ============================================
        if daily_weight_summary is not None and not daily_weight_summary.empty:
            if all(col in daily_weight_summary.columns for col in ['BM_NAV', 'MP_NAV']):
                print(f"②_누적_NAV_추이 시트 작성 중...")
                # Start 행 추가 (기준일자)
                base_date_summary = daily_weight_summary[daily_weight_summary['날짜'] == base_date]
                if base_date_summary.empty:
                    first_date = daily_weight_summary['날짜'].min()
                    base_date_summary = daily_weight_summary[daily_weight_summary['날짜'] == first_date]
                
                nav_trend = daily_weight_summary[['날짜', 'BM_NAV', 'MP_NAV']].copy()
                nav_trend.columns = ['Date', 'BM NAV', 'MP NAV']
                
                # Start 행 추가
                if not base_date_summary.empty:
                    start_row = base_date_summary.iloc[0]
                    start_df = pd.DataFrame({
                        'Date': ['Start'],
                        'BM NAV': [start_row.get('BM_NAV', 0)],
                        'MP NAV': [start_row.get('MP_NAV', 0)]
                    })
                    nav_trend = pd.concat([start_df, nav_trend], ignore_index=True)
                
                nav_trend.to_excel(
                    writer,
                    sheet_name='②_누적_NAV_추이',
                    index=False
                )
        
        # ============================================
        # ③ Active 포지션 모니터링 (절대 기준)
        # ============================================
        if '날짜' in weight_comparison_data.columns and '절대_Active_금액' in weight_comparison_data.columns:
            # Active 금액이 있는 종목만 필터링
            active_stocks = weight_comparison_data[weight_comparison_data['절대_Active_금액'] != 0].copy()
            
            if not active_stocks.empty:
                print(f"③_Active_포지션_모니터링 시트 작성 중...")
                # 각 종목별로 별도 테이블 생성
                for stock_name in active_stocks['종목명'].unique():
                    stock_data = active_stocks[active_stocks['종목명'] == stock_name].sort_values('날짜')
                    
                    active_monitoring = []
                    prev_active_amount = None
                    prev_price = None
                    total_pnl = 0.0
                    
                    # Start 행 추가 (기준일자)
                    base_date_data = stock_data[stock_data['날짜'] == base_date]
                    if base_date_data.empty:
                        first_date = stock_data['날짜'].min()
                        base_date_data = stock_data[stock_data['날짜'] == first_date]
                    
                    if not base_date_data.empty:
                        start_row = base_date_data.iloc[0]
                        start_active_amount = start_row.get('절대_Active_금액', 0)
                        start_price = start_row.get('PRICE', None)
                        
                        active_monitoring.append({
                            'Date': 'Start',
                            f'{stock_name} Active Amount': start_active_amount,
                            f'{stock_name} Return': '–',
                            f'{stock_name} Active P&L': 0.00
                        })
                        
                        prev_active_amount = start_active_amount
                        prev_price = start_price
                    
                    # 일별 데이터
                    for _, row in stock_data.iterrows():
                        date = row['날짜']
                        active_amount = row.get('절대_Active_금액', 0)
                        current_price = row.get('PRICE', None)
                        
                        # 일별 수익률 계산
                        if prev_price is not None and prev_price > 0 and current_price is not None:
                            daily_return = ((current_price / prev_price) - 1) * 100
                            # Active P&L = 전일 Active Amount × 일별 수익률
                            active_pnl = prev_active_amount * (daily_return / 100)
                            total_pnl += active_pnl
                        else:
                            daily_return = 0.0
                            active_pnl = 0.0
                        
                        active_monitoring.append({
                            'Date': date,
                            f'{stock_name} Active Amount': active_amount,
                            f'{stock_name} Return': f'{daily_return:.1f}%' if daily_return != 0 else '–',
                            f'{stock_name} Active P&L': f'{active_pnl:.3f}'
                        })
                        
                        prev_active_amount = active_amount
                        prev_price = current_price
                    
                    # 합계 행 추가
                    active_monitoring.append({
                        'Date': '합계',
                        f'{stock_name} Active Amount': '',
                        f'{stock_name} Return': '',
                        f'{stock_name} Active P&L': f'{total_pnl:.3f}'
                    })
                    
                    if active_monitoring:
                        active_df = pd.DataFrame(active_monitoring)
                        sheet_name = f'③_Active_{stock_name}' if len(active_stocks['종목명'].unique()) > 1 else '③_Active_포지션_모니터링'
                        active_df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            index=False
                        )
        
        # ============================================
        # ④ 참고용: 정규화된 비중 (보조 차트)
        # ============================================
        if '날짜' in weight_comparison_data.columns and '절대_Active_금액' in weight_comparison_data.columns:
            # Active 금액이 있는 종목만 선택
            active_stocks = weight_comparison_data[weight_comparison_data['절대_Active_금액'] != 0].copy()
            
            if not active_stocks.empty and daily_weight_summary is not None and not daily_weight_summary.empty:
                print(f"④_참고용_정규화된_비중 시트 작성 중...")
                # 각 종목별로 별도 테이블 생성
                for stock_name in active_stocks['종목명'].unique():
                    stock_data = active_stocks[active_stocks['종목명'] == stock_name].sort_values('날짜')
                    
                    normalized_weights = []
                    
                    # Start 행 추가
                    base_date_data = stock_data[stock_data['날짜'] == base_date]
                    if base_date_data.empty:
                        first_date = stock_data['날짜'].min()
                        base_date_data = stock_data[stock_data['날짜'] == first_date]
                    
                    if not base_date_data.empty:
                        start_row = base_date_data.iloc[0]
                        bm_weight = start_row.get('BM_비중', 0) * 100
                        
                        # Start일의 MP Weight 계산
                        start_date_summary = daily_weight_summary[daily_weight_summary['날짜'] == start_row['날짜']]
                        if not start_date_summary.empty and 'MP_NAV' in start_date_summary.columns:
                            mp_nav = start_date_summary.iloc[0].get('MP_NAV', 1.0)
                            mp_amount = start_row.get('MP_금액', 0)
                            if mp_nav > 0:
                                mp_weight_normalized = (mp_amount / mp_nav) * 100
                            else:
                                mp_weight_normalized = bm_weight
                        else:
                            mp_weight_normalized = bm_weight
                        
                        weight_diff = mp_weight_normalized - bm_weight
                        
                        normalized_weights.append({
                            'Date': 'Start',
                            f'{stock_name} Weight (MP, %)': f'{mp_weight_normalized:.2f}%',
                            f'{stock_name} BM 대비': f'{weight_diff:.2f}%'
                        })
                    
                    # 일별 데이터 (주요 날짜만 선택 - Start, 중간, 최종)
                    dates_sorted = sorted(stock_data['날짜'].unique())
                    # Start, 중간 1개, 최종만 선택
                    if len(dates_sorted) > 2:
                        selected_dates = [dates_sorted[0], dates_sorted[len(dates_sorted)//2], dates_sorted[-1]]
                    else:
                        selected_dates = dates_sorted
                    
                    for date in selected_dates:
                        if date == base_date or date == dates_sorted[0]:
                            continue  # Start는 이미 추가됨
                        
                        row = stock_data[stock_data['날짜'] == date].iloc[0]
                        bm_weight = row.get('BM_비중', 0) * 100
                        
                        # MP Weight (정규화) 계산
                        date_summary = daily_weight_summary[daily_weight_summary['날짜'] == date]
                        if not date_summary.empty and 'MP_NAV' in date_summary.columns:
                            mp_nav = date_summary.iloc[0].get('MP_NAV', 1.0)
                            mp_amount = row.get('MP_금액', 0)
                            if mp_nav > 0:
                                mp_weight_normalized = (mp_amount / mp_nav) * 100
                            else:
                                mp_weight_normalized = bm_weight
                        else:
                            mp_weight_normalized = bm_weight
                        
                        weight_diff = mp_weight_normalized - bm_weight
                        
                        normalized_weights.append({
                            'Date': date,
                            f'{stock_name} Weight (MP, %)': f'{mp_weight_normalized:.2f}%',
                            f'{stock_name} BM 대비': f'{weight_diff:.2f}%'
                        })
                    
                    if normalized_weights:
                        normalized_df = pd.DataFrame(normalized_weights)
                        sheet_name = f'④_정규화비중_{stock_name}' if len(active_stocks['종목명'].unique()) > 1 else '④_참고용_정규화된_비중'
                        normalized_df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            index=False
                        )
                        
                        # 주석 추가
                        from openpyxl.comments import Comment
                        ws = writer.sheets[sheet_name]
                        # 첫 번째 데이터 행에 주석 추가
                        ws.cell(row=2, column=1).comment = Comment(
                            "MP는 101% 포트이며, 본 비중은 정규화된 참고값",
                            "시스템"
                        )
        
        
        # ============================================
        # ⑤ 성과 리포트 (아주 단순하게)
        # ============================================
        if '기준일자_대비_수익률' in weight_comparison_data.columns:
            print(f"⑤_성과_리포트 시트 작성 중...")
            base_date_data = weight_comparison_data[weight_comparison_data['날짜'] == base_date].copy()
            if base_date_data.empty:
                first_date = weight_comparison_data['날짜'].min()
                base_date_data = weight_comparison_data[weight_comparison_data['날짜'] == first_date].copy()
            
            latest_date = weight_comparison_data['날짜'].max()
            latest_perf_data = weight_comparison_data[weight_comparison_data['날짜'] == latest_date].copy()
            
            if not base_date_data.empty and not latest_perf_data.empty:
                # BM 수익률 = Σ (BM_weight × 종목수익률)
                bm_return = (latest_perf_data['BM_비중'] * latest_perf_data['기준일자_대비_수익률']).sum() * 100
                
                # MP 수익률 = Σ (MP_amount × 종목수익률) / MP_NAV
                if daily_weight_summary is not None and not daily_weight_summary.empty and 'MP_NAV' in daily_weight_summary.columns:
                    latest_summary = daily_weight_summary.iloc[-1]
                    mp_nav = latest_summary.get('MP_NAV', 1.0)
                    if mp_nav > 0 and 'MP_금액' in latest_perf_data.columns:
                        # MP_amount × 종목수익률의 합
                        mp_total_return = (latest_perf_data['MP_금액'] * latest_perf_data['기준일자_대비_수익률']).sum() * 100
                        mp_return = mp_total_return / mp_nav
                    else:
                        mp_return = (latest_perf_data['전략_비중'] * latest_perf_data['기준일자_대비_수익률']).sum() * 100
                else:
                    mp_return = (latest_perf_data['전략_비중'] * latest_perf_data['기준일자_대비_수익률']).sum() * 100
                
                # Absolute Alpha (₩) = MP_NAV × (MP_return - BM_return) / 100
                if daily_weight_summary is not None and not daily_weight_summary.empty and 'MP_NAV' in daily_weight_summary.columns:
                    latest_summary = daily_weight_summary.iloc[-1]
                    mp_nav = latest_summary.get('MP_NAV', 1.0)
                    absolute_alpha = mp_nav * (mp_return - bm_return) / 100
                else:
                    absolute_alpha = (mp_return - bm_return) / 100
                
                # Relative Alpha (%) = MP_return - BM_return
                relative_alpha = mp_return - bm_return
                
                # Alpha Source 계산 (Active 금액이 있는 종목)
                alpha_source_list = []
                if '날짜' in weight_comparison_data.columns and '절대_Active_금액' in weight_comparison_data.columns:
                    latest_active = latest_perf_data[latest_perf_data['절대_Active_금액'] != 0].copy()
                    for _, row in latest_active.iterrows():
                        stock_name = row['종목명']
                        active_pct = row.get('절대_Active_비율', 0) * 100
                        if active_pct > 0:
                            alpha_source_list.append(f'{stock_name} {active_pct:.1f}% OW')
                        elif active_pct < 0:
                            alpha_source_list.append(f'{stock_name} {abs(active_pct):.1f}% UW')
                
                alpha_source = ', '.join(alpha_source_list) if alpha_source_list else 'N/A'
                
                performance_summary = pd.DataFrame({
                    '항목': ['BM 누적 수익률', 'MP 누적 수익률', 'Relative Alpha', 'Absolute Alpha', 'Alpha Source'],
                    '값': [f'{bm_return:.2f}%', f'{mp_return:.2f}%', f'{relative_alpha:.2f}%', f'{absolute_alpha:.4f} (₩ 기준)', alpha_source]
                })
                performance_summary.to_excel(
                    writer,
                    sheet_name='⑤_성과_요약',
                    index=False
                )
        
        # 전체 데이터는 제거 (핵심 정보만 제공)
        
        # 파일 저장
        writer.close()
        print(f"ExcelWriter.close() 완료")
        
        # 파일이 실제로 생성되었는지 확인
        if os.path.exists(output_path):
            file_size = os.path.getsize(output_path)
            print(f"✓ 엑셀 파일이 성공적으로 저장되었습니다!")
            print(f"  경로: {output_path}")
            print(f"  크기: {file_size:,} bytes")
        else:
            print(f"✗ 오류: 파일이 저장되지 않았습니다. 경로: {output_path}")
            return None
            
    except Exception as e:
        print(f"✗ 오류: 엑셀 파일 저장 실패 - {e}")
        import traceback
        traceback.print_exc()
        return None
    
    return output_path


if __name__ == "__main__":
    INDEX_NAME = "NDX Index"  # 지수명
    BASE_DATE = "2025-12-01"  # 기준일자 (YYYY-MM-DD)
    END_DATE = "2025-12-10"   # 종료일자 (YYYY-MM-DD)
    OUTPUT_PATH = None        # 출력 경로 (None이면 자동 생성)
    
    print(f"전략 포트폴리오 비중 검증 실행")
    print(f"=" * 50)
    print(f"지수명: {INDEX_NAME}")
    print(f"기준일자: {BASE_DATE}")
    print(f"종료일자: {END_DATE}")
    if OUTPUT_PATH:
        print(f"출력 경로: {OUTPUT_PATH}")
    else:
        print(f"출력 경로: 자동 생성")
    print(f"=" * 50)
    print()
    
    try:
        result = save_verification_excel(INDEX_NAME, BASE_DATE, END_DATE, OUTPUT_PATH)
        if result:
            print(f"\n{'=' * 50}")
            print(f"✓ 완료! 파일이 저장되었습니다.")
            print(f"{'=' * 50}")
        else:
            print(f"\n{'=' * 50}")
            print(f"✗ 실패: 파일 저장에 실패했습니다.")
            print(f"{'=' * 50}")
            sys.exit(1)
    except Exception as e:
        print(f"\n{'=' * 50}")
        print(f"✗ 오류 발생: {e}")
        print(f"{'=' * 50}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
