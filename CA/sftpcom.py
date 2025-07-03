import os
import pandas as pd
import paramiko
import win32com.client
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas.tseries.offsets import BDay
from datetime import *


class SFTPCom:
    def __init__(self, host, user, pwd, start_date, end_date):
        self.host = host
        self.user = user
        self.pwd = pwd
        self.date_list = pd.bdate_range(start_date, end_date)
        self.ssh = paramiko.SSHClient()
        self.ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    def connect_to_ftp(self):
        '''Connect to FTP. Should be executed first'''
        try:
            self.ssh.connect(hostname = self.host, username = self.user, password = self.pwd, timeout=30)
            print('Connected to ', self.host)
        except paramiko.AuthenticationException:
            print(f'[AUTH ERROR] Authentication failed for {self.host} - Check username/password')
        except paramiko.SSHException as e:
            print(f'[SSH ERROR] SSH connection failed for {self.host}: {e}')
        except Exception as e:
            print(f'[CONNECTION ERROR] Connection failed for {self.host}: {e}')

    def get_from_SNP(self, idx_code, local_path):
        # Get SPEHYDUP, SP500
        if not self.ssh.get_transport() or not self.ssh.get_transport().is_active():
            print(f'[ERROR] No active connection to {self.host}')
            return
            
        try:
            sftp = self.ssh.open_sftp()
        except Exception as e:
            print(f'[SFTP ERROR] Unable to open SFTP channel for {self.host}: {e}')
            return
            
        folder = 'WEIGHT'
        local_path2 = os.path.join(local_path, idx_code + '_' + folder)

        for dt in self.date_list:
            date_text = datetime.strftime(dt, "%Y%m%d")

            if idx_code == 'SPEHYDUP':
                file_name = date_text + '_' + idx_code + '_' + 'NCS_ADJ' + '.SDC'
                local_dir = os.path.join(local_path2, file_name)

                print('Accessing ', file_name, '...')
                ftp_dir = r'/Inbox'
                sftp.chdir(ftp_dir)
                sftp.get(file_name, local_dir)

            elif idx_code == 'SP500':
                file_name = date_text + '_' + idx_code + '_' + 'NCS_ADJ' + '.SDC'
                local_dir = os.path.join(local_path2, file_name)

                print('Accessing ', file_name, '...')
                ftp_dir = r'/Products/SP_US_EOD_NCS'
                sftp.chdir(ftp_dir)
                sftp.get(file_name, local_dir)

    def get_from_NQ_WEIGHT(self, idx_code, local_path):
        if not self.ssh.get_transport() or not self.ssh.get_transport().is_active():
            print(f'[ERROR] No active connection to {self.host}')
            return
            
        try:
            sftp = self.ssh.open_sftp()
        except Exception as e:
            print(f'[SFTP ERROR] Unable to open SFTP channel for {self.host}: {e}')
            return
            
        folder = 'WEIGHT'
        idx_code = 'IW03'
        ticker = 'NDX'

        for dt in self.date_list:
            year_text = str(dt.year)
            date_text = datetime.strftime(dt, "%m%d")

            ftp_dir1 = '/' + year_text + '/' + date_text + '/' + idx_code
            sftp.chdir(ftp_dir1)

            ftp_dir2 = ftp_dir1 + '/' + folder
            sftp.chdir(ftp_dir2)
            
            local_path2 = os.path.join(local_path, ticker + '_' + folder)

            file_list = [i for i in sftp.listdir() if i.split('_')[1] == ticker and i.split('_')[2] == 'WEOD'] # 종가 데이터만 수집
            for file in file_list:
                print('Accessing ', file, '...')
                local_dir = os.path.join(local_path2, file)
                file_name = ftp_dir2 + '/' + year_text + date_text + '_' + ticker + '_' + 'WEOD_01.txt'
                sftp.get(file_name, local_dir)

                df = pd.read_csv(local_dir, delimiter = '|')
                df.to_csv(local_dir.replace('txt', 'csv'), header = True, index = True)

                if os.path.isfile(local_dir):
                    os.remove(local_dir)

    def get_from_SX5E(self, idx_code, local_path):
        if not self.ssh.get_transport() or not self.ssh.get_transport().is_active():
            print(f'[ERROR] No active connection to {self.host}')
            return
            
        try:
            sftp = self.ssh.open_sftp()
        except Exception as e:
            print(f'[SFTP ERROR] Unable to open SFTP channel for {self.host}: {e}')
            return
            
        folder = 'WEIGHT'
        local_path2 = os.path.join(local_path, idx_code + '_' + folder)

        for dt in self.date_list:
            date_text = datetime.strftime(dt, "%Y%m%d")
            file_name = 'closecomposition' + '_' + idx_code.lower() + '_' + date_text + '.csv'
            local_dir = os.path.join(local_path2, file_name)

            if idx_code == 'SX5E':
                print('Accessing', file_name, '...')
                ftp_dir = r'/STOXX/SX5E/composition-files'
                try:
                    sftp.chdir(ftp_dir)
                    sftp.get(file_name, local_dir)

                    df = pd.read_csv(local_dir, delimiter = ';')
                    df.to_csv(local_dir.replace('closecomposition_sx5e' + '_' + date_text, date_text + '_' + idx_code + '_' + 'WEOD'), header = True, index = True) # 파일명 어제 날짜로 수정

                    if os.path.isfile(local_dir):
                        os.remove(local_dir) # 파일명 수정 후 'closecomposition_sx5e.csv' 파일 삭제
                except Exception as e:
                    if "Cannot find message" in str(e) or "No such file" in str(e) or "not a valid file path" in str(e):
                        print(f'[HOLIDAY] SX5E - {dt.strftime("%Y-%m-%d")} is a holiday, skipping...')
                    else:
                        print(f'[ERROR] SX5E processing error: {e}')

    # CA 관련 메서드들 추가
    def get_from_SNP_CA(self, idx_code, local_path):
        """S&P500, SPEHYDUP CA 데이터 다운로드"""
        sftp = self.ssh.open_sftp()
        folder = 'WEIGHT'
        local_path2 = os.path.join(local_path, idx_code + '_' + folder)
        
        # Create directory if it doesn't exist
        if not os.path.exists(local_path2):
            os.makedirs(local_path2)
            
        target_date = self.date_list[0]
        excel_dir = f'\\\\10.206.101.81\\09_idx\\해외인덱스팀\\운용관리\\CA\\{target_date.year}\\{target_date.strftime("%b")}'
        excel_path = f'{excel_dir}\\{target_date.strftime("%m%d")}.xlsx'
        if not os.path.exists(excel_dir):
            os.makedirs(excel_dir)
        if not os.path.exists(excel_path):
            wb = Workbook()
            wb.remove(wb.active)
        else:
            wb = load_workbook(excel_path)
        for dt in self.date_list:
            date_text = datetime.strftime(dt, "%Y%m%d")
            file_name3 = date_text + '_' + idx_code + '_' + 'NCS.SDE'
            temp_file = os.path.join(local_path2, 'temp_ncs_sde.txt')
            sheet_name = 'SPY' if idx_code == 'SPEHYDUP' else 'SP500'
            got_data = False
            try:
                if idx_code == 'SPEHYDUP':
                    ftp_dir = r'/Inbox'
                else:
                    ftp_dir = r'/Products/SP_US_EOD_NCS'
                sftp.chdir(ftp_dir)
                sftp.get(file_name3, temp_file)
                df = pd.read_csv(temp_file, delimiter='\t')
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    ws.delete_rows(1, ws.max_row)
                else:
                    ws = wb.create_sheet(sheet_name)
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                got_data = True
            except Exception as e:
                print(f'Error processing {file_name3} for {sheet_name}: {e}')
                # If file not found, try to copy previous available sheet
                if not got_data:
                    # Find previous available Excel file using business day logic
                    prev_date = dt - BDay(1)
                    found_prev = False
                    max_lookback = 20  # 최대 20영업일까지 뒤로 찾기
                    
                    for _ in range(max_lookback):
                        prev_excel_path = f'{excel_dir}\\{prev_date.strftime("%m%d")}.xlsx'
                        if os.path.exists(prev_excel_path):
                            try:
                                prev_wb = load_workbook(prev_excel_path)
                                if sheet_name in prev_wb.sheetnames:
                                    prev_ws = prev_wb[sheet_name]
                                    # Remove current sheet if exists
                                    if sheet_name in wb.sheetnames:
                                        ws = wb[sheet_name]
                                        wb.remove(ws)
                                    ws = wb.create_sheet(sheet_name)
                                    for row in prev_ws.iter_rows(values_only=True):
                                        ws.append(row)
                                    print(f'[HOLIDAY] Copied previous data for {sheet_name} from {prev_date.strftime("%m%d")}.xlsx')
                                    found_prev = True
                                    break
                            except Exception as e2:
                                print(f'Error copying previous sheet: {e2}')
                        prev_date = prev_date - BDay(1)
                    
                    if not found_prev:
                        print(f'[ERROR] No previous data found for {sheet_name} to copy!')
                        # Create empty sheet to prevent Excel error
                        if sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            wb.remove(ws)
                        ws = wb.create_sheet(sheet_name)
                        ws.append(['No data available'])
            finally:
                # 항상 temp 파일 삭제
                if os.path.exists(temp_file):
                    try:
                        os.remove(temp_file)
                    except Exception as e:
                        print(f'Error removing temp file {temp_file}: {e}')
        try:
            wb.save(excel_path)
            print(f'================== Excel file saved (Sheet: {sheet_name}) ==================')
        except Exception as e:
            print(f'Error saving Excel file: {e}')
        sftp.close()

    def get_from_NQ_CA(self, idx_code, local_path):
        """NASDAQ CA 데이터 다운로드"""
        sftp = self.ssh.open_sftp()
        idx_code = 'IW03'
        folder_1 = 'CAUFF'
        folder_2 = 'WEIGHT'
        ticker = 'NDX'
        local_path2 = os.path.join(local_path, ticker + '_' + folder_2)
        
        # Create directory if it doesn't exist
        if not os.path.exists(local_path2):
            os.makedirs(local_path2)
            
        target_date = self.date_list[0]
        excel_dir = f'\\\\10.206.101.81\\09_idx\\해외인덱스팀\\운용관리\\CA\\{target_date.year}\\{target_date.strftime("%b")}'
        excel_path = f'{excel_dir}\\{target_date.strftime("%m%d")}.xlsx'
        if not os.path.exists(excel_dir):
            os.makedirs(excel_dir)
        if not os.path.exists(excel_path):
            wb = Workbook()
            wb.remove(wb.active)
        else:
            wb = load_workbook(excel_path)
        sheet_name = 'NDX'
        for dt in self.date_list:
            year_text = str(dt.year)
            date_text = datetime.strftime(dt, "%m%d")
            ftp_dir1 = '/' + year_text + '/' + date_text + '/' + idx_code
            ftp_dir2 = ftp_dir1 + '/' + folder_1
            file_name = ftp_dir2 + '/' + year_text + date_text + '_' + ticker + '_' + 'CAUFF_01.txt'
            temp_file = os.path.join(local_path2, 'temp_cauff.txt')
            got_data = False
            try:
                sftp.chdir(ftp_dir2)
                sftp.get(file_name, temp_file)
                with open(temp_file, 'r', encoding='utf-8') as f:
                    content = f.read()
                lines = content.strip().split('\n')
                data_array = []
                for line in lines:
                    if line.strip():
                        fields = [field.strip() for field in line.split('|')]
                        data_array.append(fields)
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    ws.delete_rows(1, ws.max_row)
                else:
                    ws = wb.create_sheet(sheet_name)
                for row in data_array:
                    ws.append(row)
                got_data = True
            except Exception as e:
                print(f'Error processing CAUFF_01.txt for {sheet_name}: {e}')
                if not got_data:
                    # Find previous available Excel file using business day logic
                    prev_date = dt - BDay(1)
                    found_prev = False
                    max_lookback = 20  # 최대 20영업일까지 뒤로 찾기
                    
                    for _ in range(max_lookback):
                        prev_excel_path = f'{excel_dir}\\{prev_date.strftime("%m%d")}.xlsx'
                        if os.path.exists(prev_excel_path):
                            try:
                                prev_wb = load_workbook(prev_excel_path)
                                if sheet_name in prev_wb.sheetnames:
                                    prev_ws = prev_wb[sheet_name]
                                    if sheet_name in wb.sheetnames:
                                        ws = wb[sheet_name]
                                        wb.remove(ws)
                                    ws = wb.create_sheet(sheet_name)
                                    for row in prev_ws.iter_rows(values_only=True):
                                        ws.append(row)
                                    print(f'[HOLIDAY] Copied previous data for {sheet_name} from {prev_date.strftime("%m%d")}.xlsx')
                                    found_prev = True
                                    break
                            except Exception as e2:
                                print(f'Error copying previous sheet: {e2}')
                        prev_date = prev_date - BDay(1)
                    
                    if not found_prev:
                        print(f'[ERROR] No previous data found for {sheet_name} to copy!')
                        # Create empty sheet to prevent Excel error
                        if sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            wb.remove(ws)
                        ws = wb.create_sheet(sheet_name)
                        ws.append(['No data available'])
            finally:
                # 항상 temp 파일 삭제
                if os.path.exists(temp_file):
                    try:
                        os.remove(temp_file)
                    except Exception as e:
                        print(f'Error removing temp file {temp_file}: {e}')
        try:
            wb.save(excel_path)
            print(f'================== Excel file saved (Sheet: {sheet_name}) ==================')
        except Exception as e:
            print(f'Error saving Excel file: {e}')
        sftp.close()


    def get_from_SX5E_CA(self, idx_code, local_path):
        sftp = self.ssh.open_sftp()
        folder = 'WEIGHT'
        local_path2 = os.path.join(local_path, idx_code + '_' + folder)
        
        # Create directory if it doesn't exist
        if not os.path.exists(local_path2):
            os.makedirs(local_path2)
            
        target_date = self.date_list[0]
        excel_dir = f'\\\\10.206.101.81\\09_idx\\해외인덱스팀\\운용관리\\CA\\{target_date.year}\\{target_date.strftime("%b")}'
        excel_path = f'{excel_dir}\\{target_date.strftime("%m%d")}.xlsx'
        if not os.path.exists(excel_dir):
            os.makedirs(excel_dir)
        if not os.path.exists(excel_path):
            wb = Workbook()
            wb.remove(wb.active)
        else:
            wb = load_workbook(excel_path)
                
        sheet_name = 'SX5E'
        for dt in self.date_list:
            prev_date = dt - BDay(1)
            date_text = prev_date.strftime("%Y%m%d")
            file_name = f'CAForecast_sx5e_{date_text}.csv'
            temp_file = os.path.join(local_path2, 'temp_sx5e.txt')
            got_data = False
            
            try:
                ftp_dir = r'/STOXX/SX5E/additional-files'
                sftp.chdir(ftp_dir)
                sftp.get(file_name, temp_file)
                
                # 파일 내용을 읽어서 Excel에 저장 (세미콜론 구분)
                df = pd.read_csv(temp_file, delimiter=';')
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    ws.delete_rows(1, ws.max_row)
                else:
                    ws = wb.create_sheet(sheet_name)
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                got_data = True
                
            except Exception as e:
                print(f'Error processing {file_name} for {sheet_name}: {e}')
                if not got_data:
                    prev_date = dt - BDay(1)
                    found_prev = False
                    max_lookback = 20
                    
                    for _ in range(max_lookback):
                        prev_excel_path = f'{excel_dir}\\{prev_date.strftime("%m%d")}.xlsx'
                        if os.path.exists(prev_excel_path):
                            try:
                                prev_wb = load_workbook(prev_excel_path)
                                if sheet_name in prev_wb.sheetnames:
                                    prev_ws = prev_wb[sheet_name]
                                    if sheet_name in wb.sheetnames:
                                        ws = wb[sheet_name]
                                        wb.remove(ws)
                                    ws = wb.create_sheet(sheet_name)
                                    for row in prev_ws.iter_rows(values_only=True):
                                        ws.append(row)
                                    print(f'[HOLIDAY] Copied previous data for {sheet_name} from {prev_date.strftime("%m%d")}.xlsx')
                                    found_prev = True
                                    break
                            except Exception as e2:
                                print(f'Error copying previous sheet: {e2}')
                        prev_date = prev_date - BDay(1)
                    
                    if not found_prev:
                        print(f'[ERROR] No previous data found for {sheet_name} to copy!')
                        if sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            wb.remove(ws)
                        ws = wb.create_sheet(sheet_name)
                        ws.append(['No data available'])
            finally:
                if os.path.exists(temp_file):
                    try:
                        os.remove(temp_file)
                    except Exception as e:
                        print(f'Error removing temp file {temp_file}: {e}')
        
        try:
            wb.save(excel_path)
            print(f'================== Excel file saved (Sheet: {sheet_name}) ==================')
        except Exception as e:
            print(f'Error saving Excel file: {e}')
        sftp.close()


def run_sftp_download(start_date=None, end_date=None, fund_code=None, is_ca=False):
    """SFTP 다운로드 함수
    start_date: 'YYYY-MM-DD' 형식 문자열 (기본값: 어제)
    end_date: 'YYYY-MM-DD' 형식 문자열 (기본값: 어제)
    fund_code: 'SP500', 'SPEHYDUP', 'NDX', 'SX5E' 중 하나 또는 리스트 또는 None(전체)
    is_ca: True면 CA 데이터 다운로드, False면 일반 데이터 다운로드
    """
    
    try:
        # 전달받은 파라미터 사용 (None인 경우에만 기본값 설정)
        if start_date is None:
            start_date = (datetime.today() - BDay(1)).strftime('%Y-%m-%d')
        if end_date is None:
            end_date = start_date
        
        # 날짜 범위 생성 (영업일 기준)
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
        date_range = pd.bdate_range(start_dt, end_dt)
        
        if is_ca:
            folder_path = '\\\\10.206.101.81\\09_idx\\globalpassive'
        else:
            folder_path = '\\\\10.206.101.81\\09_idx\\globalpassive'
            
        print(f"[DATE] Processing date range: {start_date} to {end_date}")
        print(f"[FUND] Processing fund: {fund_code if fund_code else 'ALL'}")
        print(f"[BUSINESS DAYS] Total {len(date_range)} business days to process")
        print("=" * 60)
        
        # 각 날짜별로 개별 처리
        for current_date in date_range:
            current_date_str = current_date.strftime('%Y-%m-%d')
            
            # S&P500, S&P ESG High Yield Dividend
            if fund_code is None or (isinstance(fund_code, list) and any(f in ['SP500', 'SPEHYDUP'] for f in fund_code)) or fund_code in ['SP500', 'SPEHYDUP']:
                host = 'edx.standardandpoors.com'
                user = 'KBAs0963'
                pwd = '8zkYFO,,C'
                spftp = SFTPCom(host, user, pwd, current_date_str, current_date_str)
                spftp.connect_to_ftp()
                
                if spftp.ssh.get_transport() and spftp.ssh.get_transport().is_active():
                    if is_ca:
                        # CA 데이터 다운로드
                        if fund_code is None or (isinstance(fund_code, list) and 'SP500' in fund_code) or fund_code == 'SP500':
                            try:
                                spftp.get_from_SNP_CA('SP500', folder_path)
                            except Exception as e:
                                if "Cannot find message" in str(e) or "No such file" in str(e):
                                    print(f"[HOLIDAY] SP500 CA - {current_date_str} is a holiday, skipping...")
                                else:
                                    print(f"[ERROR] SP500 CA processing error: {e}")
                        if fund_code is None or (isinstance(fund_code, list) and 'SPEHYDUP' in fund_code) or fund_code == 'SPEHYDUP':
                            try:
                                spftp.get_from_SNP_CA('SPEHYDUP', folder_path)
                            except Exception as e:
                                if "Cannot find message" in str(e) or "No such file" in str(e):
                                    print(f"[HOLIDAY] SPEHYDUP CA - {current_date_str} is a holiday, skipping...")
                                else:
                                    print(f"[ERROR] SPEHYDUP CA processing error: {e}")
                    else:
                        # 일반 데이터 다운로드
                        if fund_code is None or (isinstance(fund_code, list) and 'SP500' in fund_code) or fund_code == 'SP500':
                            try:
                                spftp.get_from_SNP('SP500', folder_path)
                            except Exception as e:
                                if "Cannot find message" in str(e) or "No such file" in str(e):
                                    print(f"[HOLIDAY] SP500 - {current_date_str} is a holiday, skipping...")
                                else:
                                    print(f"[ERROR] SP500 processing error: {e}")
                        if fund_code is None or (isinstance(fund_code, list) and 'SPEHYDUP' in fund_code) or fund_code == 'SPEHYDUP':
                            try:
                                spftp.get_from_SNP('SPEHYDUP', folder_path)
                            except Exception as e:
                                if "Cannot find message" in str(e) or "No such file" in str(e):
                                    print(f"[HOLIDAY] SPEHYDUP - {current_date_str} is a holiday, skipping...")
                                else:
                                    print(f"[ERROR] SPEHYDUP processing error: {e}")
                else:
                    print(f"[ERROR] S&P - SFTP 연결 실패, 처리 건너뛰기")
                
                spftp.ssh.close()

            # NASDAQ
            if fund_code is None or (isinstance(fund_code, list) and 'NDX' in fund_code) or fund_code == 'NDX':
                host = 'ftp.indexes.nasdaqomx.com'
                user = 'jonghar1'
                pwd = 'Nasdaq01'
                nqftp = SFTPCom(host, user, pwd, current_date_str, current_date_str)
                nqftp.connect_to_ftp()
                
                if nqftp.ssh.get_transport() and nqftp.ssh.get_transport().is_active():
                    if is_ca:
                        # CA 데이터 다운로드
                        try:
                            nqftp.get_from_NQ_CA('NDX', folder_path)
                        except Exception as e:
                            if "Cannot find message" in str(e) or "No such file" in str(e):
                                print(f"[HOLIDAY] NDX CA - {current_date_str} is a holiday, skipping...")
                            else:
                                print(f"[ERROR] NDX CA processing error: {e}")
                    else:
                        # 일반 데이터 다운로드
                        try:
                            nqftp.get_from_NQ_WEIGHT('NDX', folder_path)
                        except Exception as e:
                            if "Cannot find message" in str(e) or "No such file" in str(e):
                                print(f"[HOLIDAY] NDX - {current_date_str} is a holiday, skipping...")
                            else:
                                print(f"[ERROR] NDX processing error: {e}")
                else:
                    print(f"[ERROR] NDX - SFTP 연결 실패, 처리 건너뛰기")
                
                nqftp.ssh.close()

            # STOXX
            if fund_code is None or (isinstance(fund_code, list) and 'SX5E' in fund_code) or fund_code == 'SX5E':
                host = 'data.stoxx.com'
                user = 'jaeminlee'
                pwd = 'Kbamquant11!'
                stftp = SFTPCom(host, user, pwd, current_date_str, current_date_str)
                stftp.connect_to_ftp()
                
                if stftp.ssh.get_transport() and stftp.ssh.get_transport().is_active():
                    if is_ca:
                        # CA 데이터 다운로드
                        try:
                            stftp.get_from_SX5E_CA('SX5E', folder_path)
                        except Exception as e:
                            if "Cannot find message" in str(e) or "No such file" in str(e):
                                print(f"[HOLIDAY] SX5E CA - {current_date_str} is a holiday, skipping...")
                            else:
                                print(f"[ERROR] SX5E CA processing error: {e}")
                    else:
                        # 일반 데이터 다운로드
                        try:
                            stftp.get_from_SX5E('SX5E', folder_path)
                        except Exception as e:
                            if "Cannot find message" in str(e) or "No such file" in str(e):
                                print(f"[HOLIDAY] SX5E - {current_date_str} is a holiday, skipping...")
                            else:
                                print(f"[ERROR] SX5E processing error: {e}")
                else:
                    print(f"[ERROR] SX5E - SFTP 연결 실패, 처리 건너뛰기")
                
                stftp.ssh.close()


        
    except Exception as e:
        print(f"[ERROR] SFTP 다운로드 중 오류 발생: {e}")
        print(f"{'='*60}\n")
        raise e